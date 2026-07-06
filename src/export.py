"""輸出模組：all.csv / filtered.csv / top_candidates.xlsx。

Excel 格式：freeze header、auto filter、URL 超連結、
total_monthly_cost 與 score 高亮、priority 上色。
"""
from __future__ import annotations

import logging
import shutil
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import LISTING_COLUMNS, Listing
from .report_html import write_html_report
from .score import to_bool, to_number
from .utils import PROJECT_ROOT

logger = logging.getLogger(__name__)

PRIORITY_FILLS = {
    "優先約看": PatternFill("solid", start_color="C6EFCE"),  # 綠
    "可備選": PatternFill("solid", start_color="FFEB9C"),    # 黃
    "先跳過": PatternFill("solid", start_color="F2F2F2"),    # 灰
}
COST_GOOD_FILL = PatternFill("solid", start_color="C6EFCE")
COST_BAD_FILL = PatternFill("solid", start_color="FFC7CE")
SCORE_FILL = PatternFill("solid", start_color="DDEBF7")
HEADER_FILL = PatternFill("solid", start_color="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True)

# Excel 只放適合人工瀏覽的欄位（去掉 raw_text 等長欄位）
XLSX_COLUMNS = [
    "priority", "score", "title", "url", "total_monthly_cost", "cost_confidence",
    "price", "management_fee", "parking_fee",
    "layout", "rooms", "size_ping", "floor",
    "life_circle_guess", "address", "community_name",
    "parking_type", "has_elevator", "can_cook",
    "location_score", "cost_score", "parking_score", "condition_score",
    "commute_score", "layout_score", "landlord_contract_score",
    "distance_to_taiyuan_note", "agent_type", "posted_time", "updated_time",
]


def _resolve(path: str | Path) -> Path:
    p = Path(path)
    return p if p.is_absolute() else PROJECT_ROOT / p


def listings_to_dataframe(listings: list[Listing]) -> pd.DataFrame:
    df = pd.DataFrame([l.to_row() for l in listings])
    return df.reindex(columns=LISTING_COLUMNS)


def write_csv(df: pd.DataFrame, path: str | Path) -> Path:
    out = _resolve(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(out, index=False, encoding="utf-8-sig")
    logger.info("已輸出 %d 筆 -> %s", len(df), out)
    return out


def apply_filter(df: pd.DataFrame, config: dict) -> pd.DataFrame:
    """filtered.csv 的條件：有車位、2-3 房、<=38000 或 unknown、非套雅房。"""
    filter_cfg = config.get("filter", {})
    allowed_rooms = set(filter_cfg.get("allowed_rooms", [2, 3]))
    max_cost = float(filter_cfg.get("max_total_monthly_cost", 38000))

    def keep(row: dict) -> bool:
        if filter_cfg.get("require_parking", True) and not to_bool(row.get("has_parking")):
            return False
        rooms = to_number(row.get("rooms"))
        if rooms is None or int(rooms) not in allowed_rooms:
            return False
        if filter_cfg.get("exclude_suite", True) and to_bool(row.get("is_suite")):
            return False
        cost = to_number(row.get("total_monthly_cost"))
        if cost is not None and cost > max_cost:
            return False
        return True

    mask = [keep(row) for row in df.to_dict(orient="records")]
    filtered = df[pd.Series(mask, index=df.index)].reset_index(drop=True)
    logger.info("篩選：%d -> %d 筆", len(df), len(filtered))
    return filtered


def write_excel(df: pd.DataFrame, path: str | Path) -> Path:
    """輸出 top_candidates.xlsx（依 score 降冪，含格式）。"""
    out = _resolve(path)
    out.parent.mkdir(parents=True, exist_ok=True)

    if "score" in df.columns:
        df = df.sort_values("score", ascending=False).reset_index(drop=True)
    columns = [c for c in XLSX_COLUMNS if c in df.columns]
    export_df = df.reindex(columns=columns)

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="top_candidates")
        ws = writer.sheets["top_candidates"]

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        col_index = {name: i + 1 for i, name in enumerate(columns)}

        for row_no in range(2, ws.max_row + 1):
            # URL 超連結
            if "url" in col_index:
                cell = ws.cell(row=row_no, column=col_index["url"])
                if cell.value:
                    cell.hyperlink = str(cell.value)
                    cell.font = Font(color="0563C1", underline="single")
            # priority 上色
            if "priority" in col_index:
                cell = ws.cell(row=row_no, column=col_index["priority"])
                fill = PRIORITY_FILLS.get(str(cell.value))
                if fill:
                    cell.fill = fill
            # total_monthly_cost 高亮（<=36000 綠、>38000 紅）
            if "total_monthly_cost" in col_index:
                cell = ws.cell(row=row_no, column=col_index["total_monthly_cost"])
                cost = to_number(cell.value)
                if cost is not None:
                    if cost <= 36000:
                        cell.fill = COST_GOOD_FILL
                    elif cost > 38000:
                        cell.fill = COST_BAD_FILL
            # score 高亮
            if "score" in col_index:
                cell = ws.cell(row=row_no, column=col_index["score"])
                cell.fill = SCORE_FILL
                cell.font = Font(bold=True)

        # 欄寬：依欄名長度給個合理值
        wide_columns = {"title": 40, "url": 45, "address": 30, "distance_to_taiyuan_note": 32}
        for name, idx in col_index.items():
            ws.column_dimensions[get_column_letter(idx)].width = wide_columns.get(name, max(12, len(name) + 2))

    logger.info("已輸出 Excel %d 筆 -> %s", len(export_df), out)
    return out


def publish_to_pages(report_path: Path, config: dict) -> Path | None:
    """把報告複製成 docs/index.html（GitHub Pages 發布目錄）。

    報告是單一自足 HTML（inline CSS/JS、外部連結只有 591 絕對網址），
    不依賴 backend / localhost / 本機路徑，直接放上靜態站就能開。
    """
    pages_dir = config.get("output", {}).get("pages_dir", "docs")
    if not pages_dir:
        return None
    docs = _resolve(pages_dir)
    docs.mkdir(parents=True, exist_ok=True)
    # 告訴 GitHub Pages 不要跑 Jekyll，原樣提供檔案
    (docs / ".nojekyll").touch()
    target = docs / "index.html"
    shutil.copyfile(report_path, target)
    logger.info("已發布報告 -> %s（GitHub Pages 目錄）", target)
    return target


def export_all(scored_df: pd.DataFrame, config: dict) -> None:
    """從 scored DataFrame 產出 filtered.csv、top_candidates.xlsx 與 report.html。"""
    output_cfg = config.get("output", {})
    filtered = apply_filter(scored_df, config)
    write_csv(filtered, output_cfg.get("filtered_csv", "output/zhubei_591_filtered.csv"))
    write_excel(scored_df, output_cfg.get("top_xlsx", "output/zhubei_591_top_candidates.xlsx"))
    report = write_html_report(scored_df, output_cfg.get("report_html", "output/zhubei_591_report.html"))
    publish_to_pages(report, config)


# ---------------------------------------------------------------------------
# State-based 輸出（availability 版）
# ---------------------------------------------------------------------------

def states_to_dataframe(states: dict) -> pd.DataFrame:
    """state -> CSV 用 DataFrame（依 STATE_COLUMNS 排序欄位）。"""
    from .state import STATE_COLUMNS
    rows = []
    for state in states.values():
        data = state.model_dump()
        data.pop("payload", None)
        rows.append(data)
    df = pd.DataFrame(rows)
    return df.reindex(columns=STATE_COLUMNS)


def states_to_report_dataframe(states: dict) -> pd.DataFrame:
    """state + payload 合併，供 HTML report 顯示完整欄位（描述、電梯等）。

    state 欄位（availability、score、priority、duplicate 等）優先於 payload。
    """
    from .state import STATE_COLUMNS
    rows = []
    for state in states.values():
        data = state.model_dump()
        payload = data.pop("payload", {}) or {}
        merged = {**payload, **{k: v for k, v in data.items() if v is not None or k not in payload}}
        rows.append(merged)
    df = pd.DataFrame(rows)
    # 確保 state 欄位齊全（payload-only 欄位保留在後面）
    ordered = [c for c in STATE_COLUMNS if c in df.columns]
    extras = [c for c in df.columns if c not in ordered]
    return df[ordered + extras]


def export_state_outputs(states: dict, config: dict) -> dict[str, Path]:
    """輸出全部 state-based CSV 與 HTML report，回傳各檔案路徑。"""
    from .availability import UNAVAILABLE_STATUSES
    from .pipeline import is_top_candidate

    output_cfg = config.get("output", {})
    df = states_to_dataframe(states)
    if not df.empty and "score" in df.columns:
        df = df.sort_values("score", ascending=False, na_position="last").reset_index(drop=True)

    paths: dict[str, Path] = {}
    paths["all_current"] = write_csv(
        df, output_cfg.get("all_current_csv", "output/zhubei_591_all_current.csv"))

    active_df = df[df["availability_status"] == "active"].reset_index(drop=True)
    paths["active"] = write_csv(
        active_df, output_cfg.get("active_csv", "output/zhubei_591_active.csv"))

    unavailable_df = df[df["availability_status"].isin(UNAVAILABLE_STATUSES)].reset_index(drop=True)
    paths["unavailable"] = write_csv(
        unavailable_df, output_cfg.get("unavailable_csv", "output/zhubei_591_unavailable.csv"))

    changes_df = df[df["status_changed"] == True].reset_index(drop=True)  # noqa: E712
    paths["status_changes"] = write_csv(
        changes_df, output_cfg.get("status_changes_csv", "output/zhubei_591_status_changes.csv"))

    top_ids = {s.listing_id for s in states.values() if is_top_candidate(s, config)}
    top_df = df[df["listing_id"].isin(top_ids)].reset_index(drop=True)
    paths["top_candidates"] = write_csv(
        top_df, output_cfg.get("top_candidates_csv", "output/zhubei_591_top_candidates.csv"))

    report_df = states_to_report_dataframe(states)
    paths["report"] = write_html_report(
        report_df, output_cfg.get("report_html", "output/zhubei_591_report.html"))
    pages = publish_to_pages(paths["report"], config)
    if pages:
        paths["pages"] = pages
    return paths
